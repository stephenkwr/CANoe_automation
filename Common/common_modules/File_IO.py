from openpyxl import *
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font
import os, sys
from datetime import datetime
from zipfile import BadZipFile

# To extract information from config file and write into output file

# to open the excel file to write into. If it does not exist, create it
def open_Output_Excel(path, test_name=None, only_if_empty=True):
    # Open or create workbook
    try:
        wb = load_workbook(path) if os.path.exists(path) else Workbook()
    except (BadZipFile, InvalidFileException):
        wb = Workbook()

    # Choose default name if none provided
    if test_name is None:
        test_name = datetime.now()

    # Write to A1 (optionally only if empty)
    ws = wb.active
    cell = ws.cell(row=1, column=1)
    if not only_if_empty or cell.value in (None, ""):
        cell.value = test_name
        if isinstance(test_name, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    return wb

# Read from a provided config file that will save each row into a list
# only require index and sound level in the config file.
# The list will contain tuple pairs where the first element is the index of the sound and the second element is the sound level to play the sound at
def read_Config_File_With_User_Input(config_file):
    if not os.path.exists(config_file):
       raise FileNotFoundError(f"Config file not found: {config_file}")
    wb = load_workbook(filename = config_file, data_only = True)
    ws = wb.active
    output = []

    for row in ws.iter_rows(min_row = 3, max_col = 2, values_only = True):
        index, sound_level = row
        if index is None or sound_level is None:
            continue
        output.append((index, sound_level))

    wb.close()
    return output

# Not as user friendly as inputs are all stored in the config file. Config file must be dropped in with executable. It must be an excel named config (config.xlsx)
# order of input is as follows in the first row. Each index (1, 2, 3, 4) is the column index
# col 1) Simulation file path (.cfg file path) - COMPULSORY
# col 2) Security key (0, 1, 2) for security access in DIAGNOSIS panel. - COMPULSORY IF NOT 0
# col 3) Number of repeats per sound (1 - 10)
# col 4) OutputExcelName
# col 5) Duration (1 or 10 seconds)
def read_Config_File_For_HW_Team():
    results = {
        "ListOfSoundToPlay" : None,
        "OutputExcelName" : None,
        "SimulationFilePath" : None,
        "Duration" : None,
        "Repeats" : None,
        "SecurityAccess" : None
    }
    config_file = ""

    if getattr(sys, 'frozen', False):
        # running as EXE
        base = os.path.dirname(sys.executable) # temp folder path
    else:
        # running as script
        base = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir)) # regular folder path


    config_file = os.path.join(base, "config.xlsx")

    if not os.path.exists(config_file):
       raise FileNotFoundError(f"Config file not found: {config_file}")
    wb = load_workbook(filename = config_file, data_only = True)
    ws = wb.active


    results["ListOfSoundToPlay"] = []
    results["SimulationFilePath"] = ws.cell(row = 2, column = 1).value
    results["SecurityAccess"] = ws.cell(row = 2, column = 2).value
    results["Repeats"] = ws.cell(row = 2, column = 3).value
    results["OutputExcelName"] = ws.cell(row = 2, column = 4).value
    # results["Duration"] = ws.cell(row = 2, column = 5).value  # if we take user input from config excel. Now hardcoded to 10s
    results["Duration"] = 12
    # print(
    #     f"{results['SimulationFilePath']=}, "
    #     f"{results['Repeats']=}, "
    #     f"{results['Duration']=}, "
    #     f"{results['OutputExcelName']=}"
    # )

    assert results["SimulationFilePath"] is not None
    
    if results["SecurityAccess"] is None:
        results["SecurityAccess"] = 0

    if results["OutputExcelName"] is None:
        results["OutputExcelName"] = "Output.xlsx"

    if results["Repeats"] is None:
        results["Repeats"] = 3

    # # For the user to choose the duration. Uncomment to choose between 1 and 10 seconds. Currently hardcoded as 10 seconds
    # if results["Duration"] != 1 and results["Duration"] > 1:
    #     results["Duration"] = 10
    # else:
    #     results["Duration"] = 10

    # to extract sounds to play (index and sound levels)
    for row in ws.iter_rows(min_row = 3, max_col = 2, values_only = True):
        index, sound_level = row
        if index is None or sound_level is None:
            continue
        results["ListOfSoundToPlay"].append((index, sound_level))

    wb.close()
    return (
        results["ListOfSoundToPlay"],
        results["OutputExcelName"],
        results["SimulationFilePath"],
        results["Duration"],
        results["Repeats"],
        results["SecurityAccess"]
    )

# write into a cell based of a wb some data with the identifying row and col
def write_Into_Cell(wb, row, col, data):
    ws = wb.active
    cell = ws.cell(row = row, column = col)
    cell.value = data

# calculate the average of a row and append the result to the end of the row
def calculate_And_Write_Average(wb, row, number_of_repeats):
    ws = wb.active
    avg = 0
    for i in range(number_of_repeats):
        avg += ws.cell(row = row, column = i + 1).value
    avg /= number_of_repeats
    write_Into_Cell(wb, row = row, col = number_of_repeats + 1, data = avg)


# simple function to bold text of selected cell
def bold_text(wb, row, col):
    bold = Font(bold=True)
    ws = wb.active
    cell = ws.cell(row = row, column = col)
    cell.font = bold
    print(f"Bolded text {cell.value}")
    
